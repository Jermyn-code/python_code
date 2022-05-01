# -*- coding:utf-8 -*-
from PIL import Image
import openpyxl
from openpyxl.styles import *
from openpyxl.utils import get_column_letter


# 程序说明:在Excel中绘制图片文件
def img2excel(image_path, excel_path):
    img = Image.open(image_path)
    sheet = "抗疫宣传画"
    wb = openpyxl.Workbook()
    ws = wb.active
    wb.create_sheet(sheet)
    ws = wb[sheet]
    # 根据imq中点的RGB,修改sheet中cel1的属性，并清空其中的数据
    for col in range(img.size[0]):
        b = get_column_letter(col + 1)
        ws.column_dimensions[b].width = 0.2
        for row in range(img.size[1]):
            ws.rowdimensions[row + 1].height = 1
            rgba = img.getpixel((col, row))
            c = ""
            for i in range(3):
                c += hex(rgba[i])[2:].zfill(2)
            file = PatternFill(fill_type="solid", fgColor=c)
            ws.cell(row + 1, col + 1).fill = file
    wb.save(excel_path)
    img.close()


if __name__ == '__main__ ':
    img_path = 'D:\\学习\\2022学期\\python_code\\Data_visualization\\R-C.jpg'
    excel_path = 'D:\\学习\\2022学期\\python_code\\Data_visualization\\covid(长春).xlsx'
    img2excel(img_path, excel_path)

