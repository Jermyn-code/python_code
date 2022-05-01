# -*- coding: utf-8 -*-
# @Author : WJY
# @Time : 2022/4/25 0025 19:50
# @Deception：绘制三种图像与一张画布上
# @Filename : 绘制三种图像.py

import matplotlib.pyplot as plt
import seaborn as sns
import pandas as pd

sns.set(font="simhei")
if __name__ == '__main__':
    fn = r"D:\学习\2022学期\python_code\Data_visualization\covid(长春).xlsx"
    plt.rcParams['font.sans-serif'] = ['SimHei']
    plt.rcParams['axes.unicode_minus'] = False
    c = "区域"
    d = pd.read_excel(fn, sheet_name="确诊病例", index_col=c)
    ax = plt.subplot(111)
    # 利用Seaborn绘制热力图
    sns.set_theme()
    sns.heatmap(d, annot=True, fmt="d", linewidths=.5, ax=ax, cmap="tab20b")
    plt.show()


import matplotlib.pyplot as plt
import seaborn as sns



def getdate(fn, col):
    d = pd.read_excel(fn, index_col=col)
    data = np.array(d.values[0:]).ravel()
    area = []
    for i in range(d.shape[0]):
        for j in range(d.shape[1]):
            area.append(d.index[i])
    area = np.array(area)
    date = []
    for i in range(d.shape[0]):
        date.append(np.array(d.columns[0:]))
    date = np.array(date).ravel()
    df = pd.DataFrame(data={"日期": date, "区域": area, "确诊病例": data})
    return df


if __name__ == '__main__':
    df = getdate(r"D:\学习\2022学期\python_code\Data_visualization\covid(长春).xlsx", "区域")
    sns.set_theme(style="white")
    sns.set_style("whitegrid", \
                  {'font.sans-serif': ['SimHei', 'SimHei']})
    g1 = sns.relplot(x="日期", y="确诊病例", hue="区域", \
                     data=df, cmap="mako", \
                     sizes=(40, 400), \
                     height=6, size="确诊病例")
    plt.show()

import pandas as pd


def getdate(fn, col):
    d = pd.read_excel(fn, index_col=col)
    data = np.array(d.values[0:]).ravel()
    area = []
    for i in range(d.shape[0]):
        for j in range(d.shape[1]):
            area.append(d.index[i])
    area = np.array(area)
    date = []
    for i in range(d.shape[0]):
        date.append(np.array(d.columns[0:]))
    date = np.array(date).ravel()
    df = pd.DataFrame(data={"日期": date, "区域": area, "确诊病例": data})
    return df


import matplotlib.pyplot as plt
import numpy as np
from mpl_toolkits.mplot3d import Axes3D

plt.rcParams['font.sans-serif'] = ['STKAITI']
plt.rcParams['axes.unicode_minus'] = False
plt.rcParams['axes.facecolor'] = '#cc00ff'
fig = plt.figure(figsize=(12, 10), facecolor='#cc00ff')
ax = Axes3D(fig)
delta = 0.125
# 生成代表X轴数据的列表
x = np.linspace(-2, 2, 10)
# 生成代表Y轴数据的列表
y = np.linspace(-2, 2, 10)
# 对x、y数据执行网格化
X, Y = np.meshgrid(x, y)

# 计算Z轴数据（高度数据）
Z = X ** 2 - Y ** 2
# 绘制3D图形
ax.plot_surface(X, Y, Z,
                rstride=1,  # rstride（row）指定行的跨度
                cstride=1,  # cstride(column)指定列的跨度
                cmap=plt.get_cmap('viridis'))  # 设置颜色映射
plt.xlabel('X轴', fontsize=15)
plt.ylabel('Y轴', fontsize=15)
ax.set_zlabel('Z轴', fontsize=15)
ax.set_title('《曲面图》', y=1.02, fontsize=25, color='white')
plt.show()
# 'twilight', 'twilight_r', 'twilight_shifted', 'twilight_shifted_r', 'viridis', 'viridis_r', 'win

from openpyxl import load_workbook
from openpyxl.drawing.image import Image


def insertimg2excel(imgPath, excelPath, x, i, y):
    imgsize = (1280 / 3, 720 / 3)  # 设置一个图像缩小的比例
    wb = load_workbook(excelPath)
    if i == 1:
        ws = wb.create_sheet('image')
    ws = wb['image']
    ws.column_dimensions[y].width = imgsize[0] * 0.14  # 修改列A的宽

    img = Image(imgPath)  # 缩放图片
    img.width, img.height = imgsize
    ws.add_image(img, x)  # 图片 插入 A1 的位置上
    ws.row_dimensions[1].height = imgsize[1] * 1  # 修改列第1行的宽

    wb.save(f'out{i}.xlsx')  # 新的结果保存输出


if __name__ == '__main__':
    for i in range(1, 4):
        imgPath = f'D:/学习/2022学期/python_code/Data_visualization/code/Figure_{i}.png'
        excelPath = r'D:\学习\2022学期\python_code\Data_visualization\covid(长春).xlsx'
        if i == 1:
            insertimg2excel(imgPath, excelPath, 'A1', i, 'A')
        if i == 2:
            excelPath = f'./out{i - 1}.xlsx'
            insertimg2excel(imgPath, excelPath, 'B1', i, 'B')
        if i == 3:
            excelPath = f'./out{i - 1}.xlsx'
            insertimg2excel(imgPath, excelPath, 'C1', i, 'C')
